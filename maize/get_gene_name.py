# -*- coding: utf-8 -*-

import os
import re
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from chromedriver_py import binary_path


excel_file = 'C:/Users/nineep/Desktop/example.xlsx'
sheet_name = 'Sheet1'
# 后台运行chrome
option = webdriver.ChromeOptions()
option.add_argument('headless')
driver = webdriver.Chrome(executable_path=binary_path, options=option)
# 加载excel
wb = load_workbook(excel_file)


def get_genes_id(sheet_name, begin_row):
    """
    从excel中获取gene_id, row num，返回字典
    """
    ws = wb[sheet_name]
    genes_id_dict = {}
    # 遍历第一列的第二行之后内容
    for row in ws.iter_rows(max_col=1, min_row=begin_row):
        for cell in row:
            gene = re.findall('Zm.+', cell.value)
            if gene:
                print('从Excel第', cell.row, '行获取的gene ID：', gene[0])
                gene_dict = {gene[0]: cell.row}
        genes_id_dict.update(gene_dict)
    return genes_id_dict


def get_gene_name_offset(sheet_name):
    ws = wb[sheet_name]
    for col in ws.iter_cols(min_col=2, max_col=2, min_row=2):
        for cell in col:
            if cell.value:
                offset = cell.row
    return offset


def insert_gene_name(sheet_name, col, row_num, text):
    ws = wb[sheet_name]
    coordinate = col + str(row_num)
    print('gene name插入Excel坐标：', coordinate)
    ws[coordinate] = text
    wb.save(excel_file)


def generate_url(sheet_name, begin_row):
    begin_url = 'https://www.maizegdb.org/search_engine/search?term='
    end_url = '&type=0&alldata=true'
    genes_id_dict = get_genes_id(sheet_name=sheet_name, begin_row=begin_row)
    urls_row_num_dict = {}
    for gene_id, row_num in genes_id_dict.items():
        url = os.path.join(begin_url + gene_id + end_url)
        url_row_num_dict = {url: row_num}
        urls_row_num_dict.update(url_row_num_dict)
    return urls_row_num_dict


def get_text(url):
    print('访问网址：', url)
    driver.get(url)
    try:
        element = WebDriverWait(driver, 240).until(
            EC.presence_of_element_located((By.ID, 'results_box')))
        text = element.get_attribute('textContent')
        no_gene_loci = re.findall('No gene loci were found matching the term', text)
        if no_gene_loci:
            print('此gene id未搜索到gene name！')
            gene_name = 'None'
        else:
            # text = driver.find_element_by_id('results_box').get_attribute('textContent')
            result_text_list = text.split()
            # 第15个字符串为gene name
            gene_name = result_text_list[14]
            print('gene name：', gene_name)
    except Exception:
        print('Error：无法找到网页的element！')
    return gene_name


def run(excel_file, sheet_name, col):
    offset = get_gene_name_offset(sheet_name=sheet_name)
    url_row_dict = generate_url(sheet_name=sheet_name, begin_row=offset)
    for url, row in url_row_dict.items():
        gene_name = get_text(url=url)
        insert_gene_name(sheet_name=sheet_name, col=col, row_num=row, text=gene_name)
    print('保存Excel文件：', excel_file)
    wb.save(excel_file)
    print('退出浏览器...')
    driver.quit()


if __name__ == '__main__':
    success = False
    while True:
        try:
            run(excel_file=excel_file, sheet_name=sheet_name, col='B')
            success = True
        except Exception:
            pass
        if success:
            break
