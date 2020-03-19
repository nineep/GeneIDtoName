# -*- coding: utf-8 -*-

import os
import re
import threading
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from chromedriver_py import binary_path


excel_file = 'C:/Users/nineep/Desktop/example.xlsx'
sheet_name = 'Sheet1'
# 后台运行chrome
option = webdriver.ChromeOptions()
option.add_argument('headless')
driver = webdriver.Chrome(executable_path=binary_path, options=option)
# 加载excel
wb = load_workbook(excel_file)


def get_genes_id(sheet_name):
    """
    从excel中获取gene_id, row num，返回字典
    """
    ws = wb[sheet_name]
    genes_id_dict = {}
    # 遍历第一列的第二行之后内容
    for row in ws.iter_rows(max_col=1, min_row=2):
        for cell in row:
            gene = re.findall('Zm.+', cell.value)
            if gene:
                print('从Excel获取的gene ID：', gene[0])
                gene_dict = {gene[0]: cell.row}
        genes_id_dict.update(gene_dict)
    return genes_id_dict


def insert_gene_name(sheet_name, col, row_num, text):
    ws = wb[sheet_name]
    coordinate = col + str(row_num)
    print('gene name插入Excel坐标：', coordinate)
    ws[coordinate] = text


def generate_url(sheet_name):
    begin_url = 'https://www.maizegdb.org/search_engine/search?term='
    end_url = '&type=0&alldata=true'
    genes_id_dict = get_genes_id(sheet_name=sheet_name)
    #urls_row_num_dict = {}
    for gene_id, row_num in genes_id_dict.items():
        url = os.path.join(begin_url + gene_id + end_url)
        #url_row_num_dict = {url: row_num}
        url_row_num_list = [url, row_num]
        yield url_row_num_list
        #urls_row_num_dict.update(url_row_num_dict)
    #return urls_row_num_dict


def get_text(url):
    print('访问网址：', url)
    driver.get(url)
    try:
        element = WebDriverWait(driver, 180).until(
            EC.presence_of_element_located((By.ID, 'results_box')))
        text = element.get_attribute('textContent')
        no_gene_loci = re.findall('No gene loci were found matching the term', text)
        if no_gene_loci:
            print('此gene id未搜索到gene name！')
            gene_name = ''
        else:
            # text = driver.find_element_by_id('results_box').get_attribute('textContent')
            result_text_list = text.split()
            # 第15个字符串为gene name
            gene_name = result_text_list[14]
            print('gene name：', gene_name)
    except Exception:
        print('Error：无法找到网页的element！')
    return gene_name


def run(sheet_name, col, row_num, url):
    gene_name = get_text(url=url)
    insert_gene_name(sheet_name=sheet_name, col=col, row_num=row_num, text=gene_name)


def multi_thread_run(excel_file, sheet_name, col, thread_num):
    generator = generate_url(sheet_name=sheet_name)
    condition = True
    while condition:
        for i in range(thread_num):
            """每开启一个线程，需要next出一个url_row_list"""
            time.sleep(1)
            try:
                url_row_list = next(generator)
                print('获得线程', i, '的参数：', url_row_list)
                url, row_num = url_row_list[0], url_row_list[1]
                print('开启线程:', i)
                thread = threading.Thread(target=run, args=(sheet_name, col, row_num, url,))
                thread.start()
                print('目前线程信息：', threading.current_thread())
            except StopIteration:
                print('需要搜索的gene id均已创建访问任务，等待任务完成...')
                condition = False
            # url_row_list = next(generator)
            # if url_row_list:
            #     print('获得线程', i, '的参数：', url_row_list)
            #     url, row_num = url_row_list[0], url_row_list[1]
            #     print('开启线程:', i)
            #     thread = threading.Thread(target=run, args=(excel_file, sheet_name, col, row_num, url,))
            #     thread.start()
            # else:
            #     print('gene id已经处理完，任务结束。')
            #     condition = False
    print('所有线程信息：', threading.enumerate())
    print('保存Excel文件：', excel_file)
    wb.save(excel_file)
    print('退出浏览器...')
    driver.quit()


if __name__ == '__main__':
    multi_thread_run(excel_file=excel_file, sheet_name=sheet_name, col='B', thread_num=4)
